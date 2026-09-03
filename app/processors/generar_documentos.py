from __future__ import annotations

from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation

DECISIONES = "PENDIENTE,APROBADO,APROBADO_HE,JUSTIFICAR_FALTA,VALIDAR_TRABAJADO,APROBAR_TURNO_Y_MEDIO,RECHAZADO"
ESTADOS = "PENDIENTE,REVISADO,CONCLUIDO"
COLUMNAS = [
    "Estado", "Empleado", "CI", "Fecha operativa", "Día", "Turno", "Área",
    "Marcaciones RAW", "Jornada reconstruida", "Horas trabajadas",
    "Retraso entrada", "Retraso comida", "Salida temprana automática",
    "Total retraso", "Excepción consolidada",
    "Impacto", "Consecuencia sin acción", "Decisión supervisor",
    "Observación supervisor",
]


def _crear_pestana(ws, filas: list[dict], nombre_tabla: str) -> None:
    ws.append(COLUMNAS)
    for fila in filas:
        ws.append([fila.get(c, "") for c in COLUMNAS])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:O{max(2, ws.max_row)}"
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="E8EAED")
        cell.font = Font(bold=True)
    if ws.max_row >= 2:
        tabla = Table(displayName=nombre_tabla, ref=f"A1:O{ws.max_row}")
        tabla.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False
        )
        ws.add_table(tabla)
    dv_estado = DataValidation(type="list", formula1=f'"{ESTADOS}"')
    dv_decision = DataValidation(type="list", formula1=f'"{DECISIONES}"')
    ws.add_data_validation(dv_estado)
    ws.add_data_validation(dv_decision)
    dv_estado.add(f"A2:A{max(2, ws.max_row)}")
    dv_decision.add(f"N2:N{max(2, ws.max_row)}")
    anchos = [14, 30, 14, 16, 13, 12, 20, 42, 28, 16, 38, 16, 25, 25, 38]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[chr(64 + i)].width = ancho


def generar_documentos_supervisor(
    incidencias: pd.DataFrame, salida: Path
) -> list[Path]:
    salida.mkdir(parents=True, exist_ok=True)
    obligatorias = {"Supervisor_Asignado", "Tipo_Personal", *COLUMNAS}
    faltantes = obligatorias - set(incidencias.columns)
    if faltantes:
        raise ValueError(f"Faltan columnas: {sorted(faltantes)}")
    archivos = []
    for supervisor, grupo in incidencias.groupby("Supervisor_Asignado", dropna=False):
        supervisor = str(supervisor or "SIN SUPERVISOR ASIGNADO").strip()
        wb = Workbook()
        wb.remove(wb.active)
        estado_col = next(
            (col for col in ("Estado_Laboral", "Estado_Empleado", "Estado empleado", "Estado_Maestro")
             if col in grupo.columns),
            None,
        )
        retirado_mask = (
            grupo[estado_col].fillna("").astype(str).str.contains("RETIR", case=False, na=False)
            if estado_col else pd.Series(False, index=grupo.index)
        )
        retirados_asignados = grupo[retirado_mask].copy()
        if not retirados_asignados.empty:
            if "Tuvo_Actividad_Mes" in retirados_asignados.columns:
                actividad = retirados_asignados["Tuvo_Actividad_Mes"].fillna(False).astype(bool)
            else:
                actividad_ci = retirados_asignados.groupby("CI")["Marcaciones RAW"].transform(
                    lambda values: values.fillna("").astype(str).str.strip().ne("").any()
                )
                actividad = actividad_ci.astype(bool)
            retirados = retirados_asignados[actividad].copy()
        else:
            retirados = retirados_asignados

        activos = grupo[~retirado_mask].copy()
        fijos = activos[~activos["Tipo_Personal"].astype(str).str.contains("Jornal", case=False, na=False)]
        jorn = activos[activos["Tipo_Personal"].astype(str).str.contains("Jornal", case=False, na=False)]
        _crear_pestana(wb.create_sheet("01_FIJOS_EVENTUALES"), fijos.to_dict("records"), "FijosEventuales")
        _crear_pestana(wb.create_sheet("02_JORNALEROS"), jorn.to_dict("records"), "Jornaleros")
        if not retirados_asignados.empty:
            _crear_pestana(
                wb.create_sheet("03_PERSONAL_RETIRADO"),
                retirados.to_dict("records"),
                "PersonalRetirado",
            )
            wb.create_sheet("04_AUDITORIA")
        else:
            wb.create_sheet("03_AUDITORIA")
        nombre = "".join(ch if ch.isalnum() else "_" for ch in supervisor).strip("_")
        destino = salida / f"Revision_PrePlanilla_{nombre}.xlsx"
        wb.save(destino)
        archivos.append(destino)
    return archivos
